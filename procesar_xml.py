import os
import xml.etree.ElementTree as ET
from openpyxl import Workbook, load_workbook

# Carpeta donde se encuentran los archivos XML
carpeta_xml = r".\Recursos\XMLSinProcesar"

# Archivo Excel donde se registrarán los datos
archivo_excel = r"C:.\Recursos\datos_xml.xlsx"

def leer_xml_y_registrar_en_excel(archivo_xml, hoja_excel):
    namespaces = {
        'cfdi': 'http://www.sat.gob.mx/cfd/3',
        'xsi': 'http://www.w3.org/2001/XMLSchema-instance',
        'pago10': 'http://www.sat.gob.mx/Pagos',
    }
    tree = ET.parse(archivo_xml)
    root = tree.getroot()

    # Extraer información del comprobante
    folio = root.get("Folio")
    fecha = root.get("Fecha")

    emisor = root.find(".//cfdi:Emisor", namespaces)
    rfc_emisor = emisor.get("Rfc")
    nombre_emisor = emisor.get("Nombre")

    conceptos = root.findall(".//cfdi:Concepto", namespaces)
    conceptos_descripcion = ", ".join([concepto.get("Descripcion") for concepto in conceptos])
    importe_total = root.get("Total")

    subtotal = root.get("SubTotal")
    iva = str(float(subtotal) * 0.16)

    datos = [
        folio,
        fecha,
        nombre_emisor,
        rfc_emisor,
        conceptos_descripcion,
        importe_total,
        iva,
        importe_total,
    ]

    # Agrega los datos en una nueva fila de la hoja de Excel
    hoja_excel.append(datos)

    # Mueve el archivo XML procesado a una carpeta diferente
    nueva_carpeta = r".\Recursos\XMLProcesados"  # Cambia la ruta según tu necesidad
    nuevo_path = os.path.join(nueva_carpeta, os.path.basename(archivo_xml))
    os.rename(archivo_xml, nuevo_path)

# Crea un archivo Excel si no existe o carga uno existente
def main():
    if not os.path.exists(archivo_excel):
        wb = Workbook()
        wb.save(archivo_excel)
    else:
        wb = load_workbook(archivo_excel)

    # Abre la primera hoja del archivo Excel
    hoja = wb.active

    # Itera a través de los archivos XML en la carpeta y registra los datos en el Excel
    for archivo in os.listdir(carpeta_xml):
        if archivo.endswith(".xml"):
            archivo_xml = os.path.join(carpeta_xml, archivo)
            leer_xml_y_registrar_en_excel(archivo_xml, hoja)

    # Guarda los cambios en el archivo Excel
    wb.save(archivo_excel)

if __name__ == "__main__":
    main()
